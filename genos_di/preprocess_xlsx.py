from __future__ import annotations

import json
import os
from pathlib import Path
import openpyxl

from collections import defaultdict
from datetime import datetime
from typing import Optional, Iterable, Any, List, Dict, Tuple

from fastapi import Request

#docling imports
from docling.backend.msexcel_backend import MsExcelDocumentBackend
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import (
    AcceleratorDevice,
    AcceleratorOptions,
    PipelineOptions
)

from docling.document_converter import (
    DocumentConverter,
    ExcelFormatOption
)
from docling.datamodel.document import ConversionResult
from docling_core.transforms.chunker import (
    BaseChunk,
    BaseChunker,
    DocChunk,
    DocMeta,
)
from docling_core.types import DoclingDocument

from pandas import DataFrame

#from docling_core.transforms.chunker import BaseChunk, BaseChunker, BaseMeta
from docling_core.types import DoclingDocument as DLDocument
from docling_core.types.doc.document import (
    #DocItem,
    DocumentOrigin,
    LevelNumber,
    ListItem,
    CodeItem,
    #SectionHeaderItem,
    #TableItem,
    #TextItem,
)
from docling_core.types.doc.labels import DocItemLabel
from docling_core.types.doc import (
    BoundingBox,
    #CoordOrigin,
    DocItemLabel,
    DoclingDocument,
    DocumentOrigin,
    #GroupLabel,
    #ImageRef,
    #ProvenanceItem,
    #Size,
    #TableCell,
    #TableData,
    #GroupItem,

    DocItem,
    PictureItem,
    SectionHeaderItem,
    TableItem,
    TextItem,
    PageItem
)
from collections import Counter
import re

import warnings
from typing import Iterable, Iterator, Optional, Union

from pydantic import BaseModel, ConfigDict, PositiveInt, TypeAdapter, model_validator
from typing_extensions import Self

try:
    import semchunk
    from transformers import AutoTokenizer, PreTrainedTokenizerBase
except ImportError:
    raise RuntimeError(
        "Module requires 'chunking' extra; to install, run: "
        "`pip install 'docling-core[chunking]'`"
    )

#from genos_utils import upload_files

# ============================================
#
# Copyright IBM Corp. 2024 - 2024
# SPDX-License-Identifier: MIT
#

"""Chunker implementation leveraging the document structure."""

class HierarchicalChunker(BaseChunker):
    r"""Chunker implementation leveraging the document layout.

    Args:
        merge_list_items (bool): Whether to merge successive list items.
            Defaults to True.
        delim (str): Delimiter to use for merging text. Defaults to "\n".
    """

    merge_list_items: bool = True

    @classmethod
    def _triplet_serialize(cls, table_df: DataFrame) -> str:

        # copy header as first row and shift all rows by one
        table_df.loc[-1] = table_df.columns  # type: ignore[call-overload]
        table_df.index = table_df.index + 1
        table_df = table_df.sort_index()

        rows = [str(item).strip() for item in table_df.iloc[:, 0].to_list()]
        cols = [str(item).strip() for item in table_df.iloc[0, :].to_list()]

        nrows = table_df.shape[0]
        ncols = table_df.shape[1]
        texts = [
            f"{rows[i]}, {cols[j]} = {str(table_df.iloc[i, j]).strip()}"
            for i in range(1, nrows)
            for j in range(1, ncols)
        ]
        output_text = ". ".join(texts)

        return output_text

    def chunk(self, dl_doc: DLDocument, **kwargs: Any) -> Iterator[BaseChunk]:
        r"""Chunk the provided document.

        Args:
            dl_doc (DLDocument): document to chunk

        Yields:
            Iterator[Chunk]: iterator over extracted chunks
        """
        heading_by_level: dict[LevelNumber, str] = {}
        list_items: list[TextItem] = []
        for item, level in dl_doc.iterate_items():
            captions = None
            if isinstance(item, DocItem):

                # first handle any merging needed
                if self.merge_list_items:
                    if isinstance(
                        item, ListItem
                    ) or (  # TODO remove when all captured as ListItem:
                        isinstance(item, TextItem)
                        and item.label == DocItemLabel.LIST_ITEM
                    ):
                        list_items.append(item)
                        continue
                    elif list_items:  # need to yield
                        yield DocChunk(
                            text=self.delim.join([i.text for i in list_items]),
                            meta=DocMeta(
                                doc_items=list_items,
                                headings=[
                                    heading_by_level[k]
                                    for k in sorted(heading_by_level)
                                ]
                                or None,
                                origin=dl_doc.origin,
                            ),
                        )
                        list_items = []  # reset

                if isinstance(item, SectionHeaderItem) or (
                    isinstance(item, TextItem)
                    and item.label in [DocItemLabel.SECTION_HEADER, DocItemLabel.TITLE]
                ):
                    level = (
                        item.level
                        if isinstance(item, SectionHeaderItem)
                        else (0 if item.label == DocItemLabel.TITLE else 1)
                    )
                    heading_by_level[level] = item.text
                    text = ''.join(str(value) for value in heading_by_level.values())

                    # remove headings of higher level as they just went out of scope
                    keys_to_del = [k for k in heading_by_level if k > level]
                    for k in keys_to_del:
                        heading_by_level.pop(k, None)
                    continue

                if (
                    isinstance(item, TextItem)
                    or ((not self.merge_list_items) and isinstance(item, ListItem))
                    or isinstance(item, CodeItem)
                ):
                    text = item.text
                elif isinstance(item, TableItem):
                    text = item.export_to_markdown(dl_doc)
                    captions = [
                        c.text for c in [r.resolve(dl_doc) for r in item.captions]
                    ] or None
                elif isinstance(item, PictureItem):
                    text = ''.join(str(value) for value in heading_by_level.values())
                else:
                    continue
                c = DocChunk(
                    text=text,
                    meta=DocMeta(
                        doc_items=[item],
                        headings=[heading_by_level[k] for k in sorted(heading_by_level)]
                        or None,
                        captions=captions,
                        origin=dl_doc.origin,
                    ),
                )
                yield c

        if self.merge_list_items and list_items:  # need to yield
            yield DocChunk(
                text=self.delim.join([i.text for i in list_items]),
                meta=DocMeta(
                    doc_items=list_items,
                    headings=[heading_by_level[k] for k in sorted(heading_by_level)]
                    or None,
                    origin=dl_doc.origin,
                ),
            )

class HybridChunker(BaseChunker):
    r"""Chunker doing tokenization-aware refinements on top of document layout chunking.

    Args:
        tokenizer: The tokenizer to use; either instantiated object or name or path of
            respective pretrained model
        max_tokens: The maximum number of tokens per chunk. If not set, limit is
            resolved from the tokenizer
        merge_peers: Whether to merge undersized chunks sharing same relevant metadata
    """

    model_config = ConfigDict(arbitrary_types_allowed=True)

    tokenizer: Union[PreTrainedTokenizerBase, str] = (
        "sentence-transformers/all-MiniLM-L6-v2"
    )
    max_tokens: int = None  # type: ignore[assignment]
    merge_peers: bool = True

    _inner_chunker: HierarchicalChunker = HierarchicalChunker()

    @model_validator(mode="after")
    def _patch_tokenizer_and_max_tokens(self) -> Self:
        self._tokenizer = (
            self.tokenizer
            if isinstance(self.tokenizer, PreTrainedTokenizerBase)
            else AutoTokenizer.from_pretrained(self.tokenizer)
        )
        if self.max_tokens is None:
            self.max_tokens = TypeAdapter(PositiveInt).validate_python(
                self._tokenizer.model_max_length
            )
        return self

    def _count_text_tokens(self, text: Optional[Union[str, list[str]]]):
        if text is None:
            return 0
        elif isinstance(text, list):
            total = 0
            for t in text:
                total += self._count_text_tokens(t)
            return total
        return len(self._tokenizer.tokenize(text))

    class _ChunkLengthInfo(BaseModel):
        total_len: int
        text_len: int
        other_len: int

    def _count_chunk_tokens(self, doc_chunk: DocChunk):
        ser_txt = self.serialize(chunk=doc_chunk)
        return len(self._tokenizer.tokenize(text=ser_txt))

    def _doc_chunk_length(self, doc_chunk: DocChunk):
        text_length = self._count_text_tokens(doc_chunk.text)
        total = self._count_chunk_tokens(doc_chunk=doc_chunk)
        return self._ChunkLengthInfo(
            total_len=total,
            text_len=text_length,
            other_len=total - text_length,
        )

    def _make_chunk_from_doc_items(
        self, doc_chunk: DocChunk, window_start: int, window_end: int
    ):
        doc_items = doc_chunk.meta.doc_items[window_start : window_end + 1]
        meta = DocMeta(
            doc_items=doc_items,
            headings=doc_chunk.meta.headings,
            captions=doc_chunk.meta.captions,
            origin=doc_chunk.meta.origin,
        )
        window_text = (
            doc_chunk.text
            if len(doc_chunk.meta.doc_items) == 1
            else self.delim.join(
                [
                    doc_item.text
                    for doc_item in doc_items
                    if isinstance(doc_item, TextItem)
                ]
            )
        )
        new_chunk = DocChunk(text=window_text, meta=meta)
        return new_chunk

    def _split_by_doc_items(self, doc_chunk: DocChunk) -> list[DocChunk]:
        chunks = []
        window_start = 0
        window_end = 0  # an inclusive index
        num_items = len(doc_chunk.meta.doc_items)
        while window_end < num_items:
            new_chunk = self._make_chunk_from_doc_items(
                doc_chunk=doc_chunk,
                window_start=window_start,
                window_end=window_end,
            )
            if self._count_chunk_tokens(doc_chunk=new_chunk) <= self.max_tokens:
                if window_end < num_items - 1:
                    window_end += 1
                    # Still room left to add more to this chunk AND still at least one
                    # item left
                    continue
                else:
                    # All the items in the window fit into the chunk and there are no
                    # other items left
                    window_end = num_items  # signalizing the last loop
            elif window_start == window_end:
                # Only one item in the window and it doesn't fit into the chunk. So
                # we'll just make it a chunk for now and it will get split in the
                # plain text splitter.
                window_end += 1
                window_start = window_end
            else:
                # Multiple items in the window but they don't fit into the chunk.
                # However, the existing items must have fit or we wouldn't have
                # gotten here. So we put everything but the last item into the chunk
                # and then start a new window INCLUDING the current window end.
                new_chunk = self._make_chunk_from_doc_items(
                    doc_chunk=doc_chunk,
                    window_start=window_start,
                    window_end=window_end - 1,
                )
                window_start = window_end
            chunks.append(new_chunk)
        return chunks

    def _split_using_plain_text(
        self,
        doc_chunk: DocChunk,
    ) -> list[DocChunk]:
        lengths = self._doc_chunk_length(doc_chunk)
        if lengths.total_len <= self.max_tokens:
            return [doc_chunk]
        else:
            # How much room is there for text after subtracting out the headers and
            # captions:
            available_length = self.max_tokens - lengths.other_len
            sem_chunker = semchunk.chunkerify(
                self._tokenizer, chunk_size=available_length
            )
            if available_length <= 0:
                warnings.warn(
                    f"Headers and captions for this chunk are longer than the total amount of size for the chunk, chunk will be ignored: {doc_chunk.text=}"  # noqa
                )
                return []
            text = doc_chunk.text
            segments = sem_chunker.chunk(text)
            chunks = [type(doc_chunk)(text=s, meta=doc_chunk.meta) for s in segments]
            return chunks

    def _merge_chunks_with_matching_metadata(self, chunks: list[DocChunk]):
        output_chunks = []
        window_start = 0
        window_end = 0  # an inclusive index
        num_chunks = len(chunks)
        while window_end < num_chunks:
            chunk = chunks[window_end]
            headings_and_captions = (chunk.meta.headings, chunk.meta.captions)
            ready_to_append = False
            if window_start == window_end:
                current_headings_and_captions = headings_and_captions
                window_end += 1
                first_chunk_of_window = chunk
            else:
                chks = chunks[window_start : window_end + 1]
                doc_items = [it for chk in chks for it in chk.meta.doc_items]
                candidate = DocChunk(
                    text=self.delim.join([chk.text for chk in chks]),
                    meta=DocMeta(
                        doc_items=doc_items,
                        headings=current_headings_and_captions[0],
                        captions=current_headings_and_captions[1],
                        origin=chunk.meta.origin,
                    ),
                )
                if (
                    headings_and_captions == current_headings_and_captions
                    and self._count_chunk_tokens(doc_chunk=candidate) <= self.max_tokens
                ):
                    # there is room to include the new chunk so add it to the window and
                    # continue
                    window_end += 1
                    new_chunk = candidate
                else:
                    ready_to_append = True
            if ready_to_append or window_end == num_chunks:
                # no more room OR the start of new metadata.  Either way, end the block
                # and use the current window_end as the start of a new block
                if window_start + 1 == window_end:
                    # just one chunk so use it as is
                    output_chunks.append(first_chunk_of_window)
                else:
                    output_chunks.append(new_chunk)
                # no need to reset window_text, etc. because that will be reset in the
                # next iteration in the if window_start == window_end block
                window_start = window_end

        return output_chunks

    def chunk(self, dl_doc: DoclingDocument, **kwargs: Any) -> Iterator[BaseChunk]:
        r"""Chunk the provided document.

        Args:
            dl_doc (DLDocument): document to chunk

        Yields:
            Iterator[Chunk]: iterator over extracted chunks
        """
        res: Iterable[DocChunk]
        res = self._inner_chunker.chunk(dl_doc=dl_doc, **kwargs)  # type: ignore
        res = [x for c in res for x in self._split_by_doc_items(c)]
        res = [x for c in res for x in self._split_using_plain_text(c)]

        if self.merge_peers:
            res = self._merge_chunks_with_matching_metadata(res)
        return iter(res)

class GenOSVectorMeta(BaseModel):
    class Config:
        extra = 'allow'

    text: str = None
    n_chars: int = None
    n_words: int = None
    n_lines: int = None
    i_page: int = None
    i_chunk_on_page: int = None
    n_chunk_of_page: int = None
    i_chunk_on_doc: int = None
    n_chunk_of_doc: int = None
    n_page: int = None
    reg_date: str = None
    chunk_bboxes: list = None
    media_files: list = None

class GenOSVectorMetaBuilder:
    def __init__(self):
        """빌더 초기화"""
        self.text: Optional[str] = None
        self.n_chars: Optional[int] = None
        self.n_words: Optional[int] = None
        self.n_lines: Optional[int] = None
        self.i_page: Optional[int] = None
        self.i_chunk_on_page: Optional[int] = None
        self.n_chunk_of_page: Optional[int] = None
        self.i_chunk_on_doc: Optional[int] = None
        self.n_chunk_of_doc: Optional[int] = None
        self.n_page: Optional[int] = None
        self.reg_date: Optional[str] = None
        self.chunk_bboxes: list = None
        self.media_files: list = None


    def set_text(self, text: str) -> "GenOSVectorMetaBuilder":
        """텍스트와 관련된 데이터를 설정"""
        self.text = text
        self.n_chars = len(text)
        self.n_words = len(text.split())
        self.n_lines = len(text.splitlines())
        return self

    def set_page_info(
            self, i_page: int, i_chunk_on_page: int, n_chunk_of_page: int
    ) -> "GenOSVectorMetaBuilder":
        """페이지 정보 설정"""
        self.i_page = i_page
        self.i_chunk_on_page = i_chunk_on_page
        self.n_chunk_of_page = n_chunk_of_page
        return self

    def set_chunk_index(self, i_chunk_on_doc: int) -> "GenOSVectorMetaBuilder":
        """문서 전체의 청크 인덱스 설정"""
        self.i_chunk_on_doc = i_chunk_on_doc
        return self

    def set_global_metadata(self, **global_metadata) -> "GenOSVectorMetaBuilder":
        """글로벌 메타데이터 병합"""
        for key, value in global_metadata.items():
            if hasattr(self, key):
                setattr(self, key, value)
        return self

    def set_chunk_bboxes(self, doc_items: list, document: DoclingDocument) -> "GenOSVectorMetaBuilder":
        self.chunk_bboxes = []
        for item in doc_items:
            for prov in item.prov:
                label = item.self_ref
                type_ = item.label
                size = document.pages.get(prov.page_no).size
                page_no = prov.page_no
                bbox = prov.bbox
                bbox_data = {'l': bbox.l / size.width,
                             't': bbox.t / size.height,
                             'r': bbox.r / size.width,
                             'b': bbox.b / size.height,
                             'coord_origin': bbox.coord_origin.value}
                self.chunk_bboxes.append({'page': page_no, 'bbox': bbox_data, 'type': type_, 'ref': label})
        return self

    def set_media_files(self, doc_items: list) -> "GenOSVectorMetaBuilder":
        temp_list = []
        for item in doc_items:
            if isinstance(item, PictureItem):
                path = str(item.image.uri)
                name = path.rsplit("/",1)[-1]
                temp_list.append({'name': name, 'type': 'image', 'ref': item.self_ref})
        self.media_files = temp_list
        return self

    def build(self) -> GenOSVectorMeta:
        """설정된 데이터를 사용해 최종적으로 GenOSVectorMeta 객체 생성"""
        return GenOSVectorMeta(
            text=self.text,
            n_chars=self.n_chars,
            n_words=self.n_words,
            n_lines=self.n_lines,
            i_page=self.i_page,
            i_chunk_on_page=self.i_chunk_on_page,
            n_chunk_of_page=self.n_chunk_of_page,
            i_chunk_on_doc=self.i_chunk_on_doc,
            n_chunk_of_doc=self.n_chunk_of_doc,
            n_page=self.n_page,
            reg_date=self.reg_date,
            chunk_bboxes=self.chunk_bboxes,
            media_files=self.media_files,
        )


class DocumentProcessor:

    def __init__(self):
        self.page_chunk_counts = defaultdict(int)
        device = AcceleratorDevice.AUTO
        num_threads = 4
        accelerator_options = AcceleratorOptions(num_threads=num_threads, device=device)
        simple_pipeline_options = PipelineOptions()
        simple_pipeline_options.accelerator_options = accelerator_options

        self.converter = DocumentConverter(
            format_options={
                InputFormat.XLSX: ExcelFormatOption(
                    pipeline_options=simple_pipeline_options,
                    backend=MsExcelDocumentBackend
                )
            }
        )

    def load_documents_with_docling(self, file_path: str, **kwargs: dict) -> DoclingDocument:
        conv_result: ConversionResult = self.converter.convert(file_path, raises_on_error=False)
        return conv_result.document

    def load_documents(self, file_path: str, **kwargs: dict) -> DoclingDocument:
        return self.load_documents_with_docling(file_path, **kwargs)

    def split_documents(self, documents: DoclingDocument, **kwargs: dict) -> List[DocChunk]:
        chunker: HybridChunker = HybridChunker()

        chunks: List[DocChunk] = list(chunker.chunk(dl_doc=documents, **kwargs))
        return chunks

    def safe_join(self, iterable):
        if not isinstance(iterable, (list, tuple, set)):
            return ''
        return ''.join(map(str, iterable)) + '\n'

    def compose_vectors(self, document: DoclingDocument, chunks: List[DocChunk], file_path: str, **kwargs: dict) -> \
            list[dict]:
        global_metadata = dict(
            n_chunk_of_doc=len(chunks),
            n_page=document.num_pages(),
            reg_date=datetime.now().isoformat(timespec='seconds') + 'Z'
        )

        chunk_index_on_page = 0
        vectors = []
        for chunk_idx, chunk in enumerate(chunks):
            content = self.safe_join(chunk.meta.headings) + chunk.text

            vector = (GenOSVectorMetaBuilder()
                    .set_text(content)
                    .set_page_info(1, chunk_idx, len(chunks))
                    .set_chunk_index(chunk_idx)
                    .set_global_metadata(**global_metadata)
                    .set_chunk_bboxes(chunk.meta.doc_items, document)
                    .set_media_files(chunk.meta.doc_items)
                    ).build()
            vectors.append(vector)
            chunk_index_on_page += 1

        return vectors
    """
    def get_media_files(self, doc_items: list):
        temp_list = []
        for item in doc_items:
            if isinstance(item, PictureItem):
                path = str(item.image.uri)
                name = path.rsplit("/",1)[-1]
                temp_list.append({ 'path': path, 'name': name})
        return temp_list
    """
    async def __call__(self, request: Request, file_path: str, **kwargs: dict):
        output_path, output_file = os.path.split(file_path)
        filename, _ = os.path.splitext(output_file)
        artifacts_dir = Path(f"{output_path}/{filename}")
        if artifacts_dir.is_absolute():
            reference_path = None
        else:
            reference_path = artifacts_dir.parent

        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 시트명 반복
        chunks: List[DocChunk] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 새로운 워크북 생성
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name

            # 기존 시트의 셀 내용을 새 워크북으로 복사
            for row in ws.iter_rows(values_only=True):
                new_ws.append(row)
            # 각 시트별로 저장 (sheet_name을 파일명으로 활용)
            sheet_path = f"{output_path}/{sheet_name}.xlsx"
            new_wb.save(sheet_path)

            document: DoclingDocument = self.load_documents(sheet_path, **kwargs)
            os.remove(sheet_path)
            # await assert_cancelled(request)
            document = document._with_pictures_refs(image_dir=Path(f"{artifacts_dir}/{sheet_name}"), reference_path=Path(f"{reference_path}/{sheet_name}"))

            # Extract Chunk from DoclingDocument
            chunk: List[DocChunk] = self.split_documents(document, **kwargs)
            # await assert_cancelled(request)
            chunks.extend(chunk)

        vectors = []
        if len(chunks) > 1:
            vectors: list[dict] = self.compose_vectors(document, chunks, file_path, **kwargs)
        else:
            raise GenosServiceException(1, f"chunk length is 0")

        """
        # 미디어 파일 업로드 방법
        media_files = [
            { 'path': '/tmp/graph.jpg', 'name': 'graph.jpg', 'type': 'image' },
            { 'path': '/result/1/graph.jpg', 'name': '1/graph.jpg', 'type': 'image' },
        ]

        # 업로드 요청 시에는 path, name 필요
        file_list = [{k: v for k, v in file.items() if k != 'type'} for file in media_files]
        await upload_files(file_list, request=request)

        # 메타에 저장시에는 name, type 필요
        meta = [{k: v for k, v in file.items() if k != 'path'} for file in media_files]
        vectors[0].media_files = meta
        """

        return vectors


class GenosServiceException(Exception):
    # GenOS 와의 의존성 부분 제거를 위해 추가
    def __init__(self, error_code: str, error_msg: Optional[str] = None, msg_params: Optional[dict] = None) -> None:
        self.code = 1
        self.error_code = error_code
        self.error_msg = error_msg or "GenOS Service Exception"
        self.msg_params = msg_params or {}

    def __repr__(self) -> str:
        class_name = self.__class__.__name__
        return f"{class_name}(code={self.code!r}, errMsg={self.error_msg!r})"


# GenOS 와의 의존성 제거를 위해 추가
async def assert_cancelled(request: Request):
    if await request.is_disconnected():
        raise GenosServiceException(1, f"Cancelled")
